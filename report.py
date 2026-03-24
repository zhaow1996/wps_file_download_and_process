import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import warnings
import requests
import json
import time
import hmac
import hashlib
import base64
import urllib.parse
from datetime import datetime
from azure.storage.blob import BlobServiceClient, BlobClient, ContainerClient
import logging
import sys
import io
import configparser

warnings.filterwarnings('ignore')


# 配置日志记录
def setup_logging():
    # 创建logger
    logger = logging.getLogger()
    logger.setLevel(logging.INFO)

    # 清除现有的处理器
    logger.handlers = []

    # 创建控制台处理器
    console_handler = logging.StreamHandler()
    console_handler.setLevel(logging.INFO)

    # 创建文件处理器
    file_handler = logging.FileHandler('excel_merge.log', encoding='utf-8')
    file_handler.setLevel(logging.INFO)

    # 创建格式器
    formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
    console_handler.setFormatter(formatter)
    file_handler.setFormatter(formatter)

    # 添加处理器到logger
    logger.addHandler(console_handler)
    logger.addHandler(file_handler)

    return logger


logger = setup_logging()

# 钉钉机器人配置
DINGTALK_WEBHOOK = "https://oapi.dingtalk.com/robot/send?access_token=e551d587af7d30cb4fd9a92bd52cd886450d84dd0a1ef61922efc7cf81cb297b"  # 钉钉机器人的Webhook地址
DINGTALK_SECRET = "SECba5744b2a1f21ce654ea60ecbef4c09af37437f6a73fb628f6028162e4197850"  # 钉钉机器人的签名密钥
AT_MOBILES = ["18504272160"]  # 需要@的手机号列表
AT_ALL = False  # 是否@所有人

# 库存数据的特定列名
INVENTORY_COLUMNS = [
    "物料编码", "物料名称", "物料属性", "存储容量", "所属基地",
    "所属厂区", "无法正常出库量（kg）", "上周末库存（kg）",
    "实时库存（除去无法正常出库量kg）", "仓库最大存放量（kg）",
    "入库 星期一", "出库 星期一", "入库 星期二", "出库 星期二",
    "入库 星期三", "出库 星期三", "入库 星期四", "出库 星期四",
    "入库 星期五", "出库 星期五", "入库 星期六", "出库 星期六",
    "入库 星期日", "出库 星期日", "本周总计入库(kg)", "本周总计出库(kg)"
]

# 采购数据的列名
PURCHASE_COLUMNS = [
    "物料编码", "物料名称", "所属厂区", "采购负责人",
    "采购策略审批人", "协议类型", "支付方式", "供应商",
    "合同价格", "当日价格", "签约合同量", "合同已执行量",
    "待执行量", "关注状态", "价格及供应趋势判断", "所属基地"
]


def send_dingtalk_message(webhook, secret, message, title="Excel合并程序通知"):
    """
    发送钉钉机器人消息（支持签名认证）
    """
    if not webhook or not secret:
        logger.warning("钉钉机器人配置不完整，跳过消息发送")
        logger.info(f"标题: {title}")
        logger.info(f"内容: {message}")
        return False

    try:
        # 生成签名
        timestamp = str(round(time.time() * 1000))
        string_to_sign = f'{timestamp}\n{secret}'

        hmac_code = hmac.new(
            secret.encode('utf-8'),
            string_to_sign.encode('utf-8'),
            digestmod=hashlib.sha256
        ).digest()

        sign = urllib.parse.quote_plus(base64.b64encode(hmac_code))

        # 构建webhook URL
        webhook_url = f"{webhook}&timestamp={timestamp}&sign={sign}"

        # 构建消息体
        data = {
            "msgtype": "markdown",
            "markdown": {
                "title": title,
                "text": f"## {title}\n\n{message}"
            },
            "at": {
                "atMobiles": AT_MOBILES if AT_MOBILES else [],
                "isAtAll": AT_ALL
            }
        }

        headers = {
            'Content-Type': 'application/json',
            'Charset': 'UTF-8'
        }

        logger.info(f"发送钉钉消息: {title}")
        response = requests.post(
            webhook_url,
            data=json.dumps(data, ensure_ascii=False).encode('utf-8'),
            headers=headers,
            timeout=10
        )

        if response.status_code == 200:
            result = response.json()
            if result.get('errcode') == 0:
                logger.info("钉钉消息发送成功")
                return True
            else:
                error_msg = result.get('errmsg', '未知错误')
                logger.error(f"钉钉消息发送失败: {error_msg}")
                return False
        else:
            logger.error(f"钉钉消息发送失败，HTTP状态码: {response.status_code}")
            logger.error(f"响应内容: {response.text}")
            return False

    except Exception as e:
        logger.error(f"发送钉钉消息时出错: {str(e)}")
        return False


def send_status_update(operation, status, message, details=None, webhook=None, secret=None):
    """
    发送状态更新消息到钉钉
    """
    current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    if not webhook or not secret:
        logger.warning("钉钉配置不完整，跳过消息发送")
        logger.info(f"状态: {status}, 操作: {operation}, 消息: {message}")
        return False

    # if status == "start":
    #     title = f"Excel合并程序启动"
    #     content = f"**时间**: {current_time}\n\n**操作**: {operation}\n\n**状态**: 开始执行"

    if status == "success":
        title = f"Excel合并完成"
        content = f"**时间**: {current_time}\n\n**操作**: {operation}\n\n**状态**: 成功\n\n**详情**: {message}"

    elif status == "warning":
        title = f"Excel合并警告"
        content = f"**时间**: {current_time}\n\n**操作**: {operation}\n\n**状态**: 警告\n\n**详情**: {message}"

    elif status == "error":
        title = f"Excel合并错误"
        content = f"**时间**: {current_time}\n\n**操作**: {operation}\n\n**状态**: 失败\n\n**详情**: {message}"

    elif status == "duplicates":
        title = f"发现重复物料"
        content = f"**时间**: {current_time}\n\n**操作**: {operation}\n\n**状态**: 发现重复物料\n\n**详情**: {message}"
        if details:
            content += f"\n\n**重复物料详情**:\n"
            for item in details:
                content += f"- {item['物料名称']} (出现{item['次数']}次)\n"
            content += f"\n**处理方式**: 已自动删除重复项，保留首次出现"

    else:
        title = f"Excel合并状态更新"
        content = f"**时间**: {current_time}\n**操作**: {operation}\n**状态**: {status}\n**详情**: {message}"

    return send_dingtalk_message(webhook, secret, content, title)


def get_dingtalk_config():
    """
    获取钉钉配置信息
    """
    # 首先尝试从环境变量获取
    webhook = os.environ.get('DINGTALK_WEBHOOK', DINGTALK_WEBHOOK)
    secret = os.environ.get('DINGTALK_SECRET', DINGTALK_SECRET)

    # 如果环境变量中没有，尝试从配置文件读取
    config_path = os.path.join(os.path.dirname(__file__), 'config.ini')
    if os.path.exists(config_path):
        try:
            config = configparser.ConfigParser()
            config.read(config_path, encoding='utf-8')

            if 'dingtalk' in config:
                webhook = config['dingtalk'].get('webhook', webhook)
                secret = config['dingtalk'].get('secret', secret)
        except Exception as e:
            logger.error(f"读取配置文件失败: {e}")

    return webhook, secret


def get_azure_config():
    """
    获取Azure配置信息
    """
    connection_string = os.environ.get('AZURE_CONNECTION_STRING',
                                       "DefaultEndpointsProtocol=https;AccountName=zwlalalala;AccountKey=43h4Kbt6xg3KlXzsfYUsAR90CQAAw2uL8CETaMWBRaaRcFoNaLGQNhJXyundxdIH6Ud4Gn7Ncrme+ASts9Mttg==;EndpointSuffix=core.windows.net")
    container_name = os.environ.get('AZURE_CONTAINER_NAME', "excel")

    config_path = os.path.join(os.path.dirname(__file__), 'config.ini')
    if os.path.exists(config_path):
        try:
            config = configparser.ConfigParser()
            config.read(config_path, encoding='utf-8')

            if 'azure' in config:
                connection_string = config['azure'].get('connection_string', connection_string)
                container_name = config['azure'].get('container_name', container_name)
        except Exception as e:
            logger.error(f"读取配置文件失败: {e}")

    return connection_string, container_name


# 获取配置
DINGTALK_WEBHOOK, DINGTALK_SECRET = get_dingtalk_config()


def filter_empty_material_rows(df):
    """
    过滤掉第二列（物料名称）为空的行
    """
    if df.empty or len(df.columns) < 2:
        return df

    # 获取第二列（物料名称列）
    material_col = df.iloc[:, 1]

    # 过滤掉物料名称为空的行
    filtered_df = df[~material_col.isna()]

    return filtered_df


def process_inventory_data(df, expected_columns=26):
    """
    处理库存数据，重命名列并确保有足够的列
    """
    # 如果列数不足，添加缺失的列
    if len(df.columns) < expected_columns:
        for i in range(len(df.columns), expected_columns):
            df[f'临时列_{i}'] = ""

    # 只保留所需列数
    df = df.iloc[:, :expected_columns]

    # 重命名列
    df.columns = INVENTORY_COLUMNS[:len(df.columns)]

    return df


def process_purchase_data(df, expected_columns=16):
    """
    处理采购数据，确保有16列并按固定列名重命名
    """
    # 如果列数不足，添加缺失的列
    if len(df.columns) < expected_columns:
        for i in range(len(df.columns), expected_columns):
            df[f'临时列_{i}'] = ""

    # 只保留所需列数
    df = df.iloc[:, :expected_columns]

    # 重命名列
    df.columns = PURCHASE_COLUMNS[:len(df.columns)]

    # 根据采购负责人填写所属基地
    if len(df.columns) > 3:
        def get_base(purchaser):
            if pd.isna(purchaser):
                return "上虞基地"
            purchaser_str = str(purchaser).strip()
            if purchaser_str in ["王群", "任文顺", "孙鑫荣"]:
                return "山东基地"
            else:
                return "上虞基地"

        if len(df.columns) > 15:
            df.iloc[:, 15] = df.iloc[:, 3].apply(get_base)

    return df


def check_duplicate_material_names(df, sheet_name, purchase_df=None):
    """
    检查物料名称列（第二列）的重复值
    如果提供了采购数据，还会查找对应的采购负责人
    """
    duplicate_list = []

    if df.empty:
        logger.info(f"{sheet_name} 数据为空，无法检查重复物料")
        return duplicate_list

    if len(df.columns) < 2:
        logger.warning(f"{sheet_name} 列数不足，无法检查第二列")
        return duplicate_list

    # 获取第二列（物料名称列）
    material_col = df.iloc[:, 1]

    if material_col.isna().all():
        logger.info(f"{sheet_name} 第二列全为空，无法检查重复物料")
        return duplicate_list

    # 查找重复值
    duplicates = material_col[material_col.duplicated(keep=False)]

    if duplicates.empty:
        logger.info(f"{sheet_name} 中没有发现重复的物料名称")
        return duplicate_list
    else:
        # 获取唯一的重复物料名称
        unique_duplicates = duplicates.unique()
        logger.info(f"{sheet_name} 中发现 {len(unique_duplicates)} 个重复的物料名称:")

        # 创建物料名称到采购负责人的映射
        purchaser_map = {}
        if purchase_df is not None and not purchase_df.empty and len(purchase_df.columns) >= 4:
            for _, row in purchase_df.iterrows():
                material_name = row.iloc[1] if len(row) > 1 else None
                purchaser = row.iloc[3] if len(row) > 3 else None
                if pd.notna(material_name) and pd.notna(purchaser):
                    purchaser_map[str(material_name).strip()] = str(purchaser).strip()

        for material in unique_duplicates:
            if pd.isna(material):
                continue

            count = (material_col == material).sum()
            duplicate_rows = material_col[material_col == material].index.tolist()
            excel_rows = [r + 2 for r in duplicate_rows]  # 转换为Excel行号

            # 获取采购负责人
            purchaser = purchaser_map.get(str(material).strip(), "未知")

            duplicate_info = {
                "物料名称": str(material),
                "次数": int(count),
                "行号": excel_rows,
                "采购负责人": purchaser
            }
            duplicate_list.append(duplicate_info)

            logger.info(f"  - '{material}': 出现 {count} 次，位于行: {excel_rows}, 采购负责人: {purchaser}")

        return duplicate_list


def remove_duplicate_material_rows(df, duplicate_list):
    """
    根据重复物料列表删除重复行（保留第一次出现的行）
    """
    if not duplicate_list or df.empty:
        return df, 0

    original_count = len(df)

    # 获取所有重复的物料名称
    duplicate_materials = [item['物料名称'] for item in duplicate_list]

    # 分离重复行和非重复行
    mask = df.iloc[:, 1].isin(duplicate_materials)
    non_duplicate_df = df[~mask]  # 非重复行

    # 对于每个重复物料，只保留第一次出现
    cleaned_duplicates = []
    for material in duplicate_materials:
        material_rows = df[df.iloc[:, 1] == material]
        if not material_rows.empty:
            cleaned_duplicates.append(material_rows.iloc[0:1])  # 只保留第一行

    # 合并非重复行和清理后的重复行
    if cleaned_duplicates:
        df_cleaned = pd.concat([non_duplicate_df] + cleaned_duplicates, ignore_index=True)
    else:
        df_cleaned = non_duplicate_df.copy()

    deleted_count = original_count - len(df_cleaned)

    if deleted_count > 0:
        logger.info(f"已删除 {deleted_count} 行重复物料数据")
        logger.info(f"保留的重复物料: {duplicate_materials}")

    return df_cleaned.reset_index(drop=True), deleted_count


def check_numeric_columns(df, sheet_name, start_col, end_col, start_row=1):
    """
    检查指定列范围内是否存在非数字单元格（空值不算）
    """
    if df.empty:
        logger.info(f"{sheet_name} 数据为空，无法检查数字格式")
        return []

    if len(df.columns) <= end_col:
        logger.warning(f"{sheet_name} 列数不足，无法检查指定列范围")
        return []

    non_numeric_cells = []

    # 遍历指定列范围
    for col_idx in range(start_col, end_col + 1):
        col_name = df.columns[col_idx]

        # 遍历指定行范围（跳过标题行）
        for row_idx in range(start_row, len(df)):
            cell_value = df.iloc[row_idx, col_idx]

            # 跳过空值
            if pd.isna(cell_value) or cell_value == "":
                continue

            # 检查是否为数字
            try:
                float(cell_value)
            except (ValueError, TypeError):
                excel_row = row_idx + 2
                non_numeric_cells.append({
                    '行号': excel_row,
                    '列名': col_name,
                    '值': str(cell_value)[:50]  # 只取前50个字符
                })

    return non_numeric_cells


def delete_specific_inventory_rows(df):
    """
    删除库存数据中特定的行
    """
    if df.empty:
        return df

    original_count = len(df)

    delete_condition = (
            ((df.iloc[:, 1] == "碳酸二甲酯") & (df.iloc[:, 4] == "上虞基地")) |
            ((df.iloc[:, 1] == "甲醇") & (df.iloc[:, 4] == "上虞基地")) |
            ((df.iloc[:, 1] == "丙酮") & (df.iloc[:, 4] == "上虞基地")) |
            ((df.iloc[:, 1] == "N,N-二甲基甲酰胺") & (df.iloc[:, 4] == "上虞基地")) |
            ((df.iloc[:, 1] == "无水乙醇") & (df.iloc[:, 4] == "上虞基地")) |
            ((df.iloc[:, 1] == "氯化亚砜") & (df.iloc[:, 4] == "上虞基地")) |
            ((df.iloc[:, 1] == "68%哌嗪") & (df.iloc[:, 4] == "上虞基地"))
    )

    df = df[~delete_condition]

    deleted_count = original_count - len(df)
    if deleted_count > 0:
        logger.info(f"已删除库存数据中 {deleted_count} 行特定数据")

    return df.reset_index(drop=True)


def upload_to_azure_blob(local_file_path, connection_string, container_name, blob_name=None):
    """
    将本地文件上传到Azure Blob Storage
    """
    if not os.path.exists(local_file_path):
        error_msg = f"本地文件不存在: {local_file_path}"
        logger.error(error_msg)
        return False, error_msg

    if blob_name is None:
        blob_name = os.path.basename(local_file_path)

    try:
        blob_service_client = BlobServiceClient.from_connection_string(connection_string)
        container_client = blob_service_client.get_container_client(container_name)

        if not container_client.exists():
            logger.info(f"容器 '{container_name}' 不存在，正在创建...")
            container_client = blob_service_client.create_container(container_name)
            logger.info(f"容器 '{container_name}' 创建成功")

        blob_client = blob_service_client.get_blob_client(
            container=container_name,
            blob=blob_name
        )

        logger.info(f"正在上传文件到Azure Blob: {local_file_path} -> {container_name}/{blob_name}")

        with open(local_file_path, "rb") as data:
            blob_client.upload_blob(data, overwrite=True)

        blob_url = blob_client.url

        logger.info(f"文件上传成功!")
        logger.info(f"   Blob URL: {blob_url}")

        return True, blob_url

    except Exception as e:
        error_msg = f"上传文件到Azure Blob时出错: {str(e)}"
        logger.error(f"{error_msg}")
        return False, error_msg


def merge_excel_sheets(folder_path, file1_name, file2_name, output_filename):
    """
    合并两个Excel文件中的特定sheet
    """
    try:
        file1_path = os.path.join(folder_path, file1_name)
        file2_path = os.path.join(folder_path, file2_name)
        output_path = os.path.join(folder_path, output_filename)

        if not os.path.exists(file1_path):
            error_msg = f"文件 {file1_path} 不存在"
            logger.error(error_msg)
            return False, error_msg, None, None

        if not os.path.exists(file2_path):
            error_msg = f"文件 {file2_path} 不存在"
            logger.error(error_msg)
            return False, error_msg, None, None

        # 获取文件最后修改时间
        try:
            file1_time = datetime.fromtimestamp(os.path.getmtime(file1_path))
            file2_time = datetime.fromtimestamp(os.path.getmtime(file2_path))
            time_format = "%Y-%m-%d %H:%M:%S"
            file1_time_str = file1_time.strftime(time_format)
            file2_time_str = file2_time.strftime(time_format)
        except Exception as e:
            logger.error(f"获取文件修改时间时出错: {e}")
            file1_time_str = "未知"
            file2_time_str = "未知"

        if os.path.exists(output_path):
            logger.info(f"删除已存在的输出文件: {output_filename}")
            os.remove(output_path)

        xls1 = pd.ExcelFile(file1_path)
        xls2 = pd.ExcelFile(file2_path)

        sheets1 = xls1.sheet_names
        sheets2 = xls2.sheet_names

        logger.info(f"文件1 '{file1_name}' 中的sheet: {sheets1}")
        logger.info(f"文件2 '{file2_name}' 中的sheet: {sheets2}")

        inventory_sheets = []
        purchase_sheets = []

        for sheet in sheets1:
            if "库存数据" in sheet:
                inventory_sheets.append((file1_path, sheet))
            else:
                purchase_sheets.append((file1_path, sheet))

        for sheet in sheets2:
            if "库存数据" in sheet:
                inventory_sheets.append((file2_path, sheet))
            else:
                purchase_sheets.append((file2_path, sheet))

        logger.info(f"找到库存数据sheet: {[s[1] for s in inventory_sheets]}")
        logger.info(f"找到采购数据sheet: {[s[1] for s in purchase_sheets]}")

        duplicate_info = []
        all_non_numeric_cells = []

        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:

            # 1. 处理库存数据
            logger.info("正在合并库存数据...")
            inventory_combined = []
            for file_path, sheet_name in inventory_sheets:
                try:
                    # 库存数据从第三行开始读取（跳过前两行标题）
                    df = pd.read_excel(file_path, sheet_name=sheet_name, header=1)
                    df = filter_empty_material_rows(df)
                    df = process_inventory_data(df, len(INVENTORY_COLUMNS))
                    inventory_combined.append(df)
                    logger.info(f"  - 已添加: {sheet_name} (共{len(df)}行)")
                except Exception as e:
                    logger.error(f"  - 读取 {sheet_name} 时出错: {e}")

            if inventory_combined:
                final_inventory = pd.concat(inventory_combined, ignore_index=True)
                final_inventory = delete_specific_inventory_rows(final_inventory)
                final_inventory.to_excel(writer, sheet_name='库存数据', index=False)
                logger.info(f"库存数据合并完成，总计 {len(final_inventory)} 行")
            else:
                final_inventory = pd.DataFrame(columns=INVENTORY_COLUMNS)
                final_inventory.to_excel(writer, sheet_name='库存数据', index=False)
                logger.warning("未找到库存数据sheet，已创建空sheet")

            # 2. 处理采购数据
            logger.info("正在合并采购数据...")
            purchase_combined = []
            for file_path, sheet_name in purchase_sheets:
                try:
                    # 采购数据从第二行开始读取（跳过第一行标题）
                    df = pd.read_excel(file_path, sheet_name=sheet_name, header=0)
                    df = filter_empty_material_rows(df)
                    df = process_purchase_data(df, len(PURCHASE_COLUMNS))
                    purchase_combined.append(df)
                    logger.info(f"  - 已添加: {sheet_name} (共{len(df)}行)")
                except Exception as e:
                    logger.error(f"  - 读取 {sheet_name} 时出错: {e}")

            if purchase_combined:
                final_purchase = pd.concat(purchase_combined, ignore_index=True)
                final_purchase.to_excel(writer, sheet_name='采购数据', index=False)
                logger.info(f"采购数据合并完成，总计 {len(final_purchase)} 行")

                # 检查库存数据中的重复物料，并关联采购负责人
                if not final_inventory.empty:
                    duplicate_info = check_duplicate_material_names(
                        final_inventory,
                        "库存数据",
                        final_purchase
                    )
            else:
                final_purchase = pd.DataFrame(columns=PURCHASE_COLUMNS)
                final_purchase.to_excel(writer, sheet_name='采购数据', index=False)
                logger.warning("未找到采购数据sheet，已创建空sheet")

        # 检查数字格式
        logger.info("正在检查数字格式...")
        try:
            merged_file = pd.ExcelFile(output_path, engine='openpyxl')  # 指定engine

            if '采购数据' in merged_file.sheet_names:
                purchase_df = pd.read_excel(output_path, sheet_name='采购数据', engine='openpyxl')
                non_numeric = check_numeric_columns(purchase_df, "采购数据", 8, 12)
                all_non_numeric_cells.extend(non_numeric)
                if non_numeric:
                    logger.warning(f"采购数据中发现 {len(non_numeric)} 个非数字单元格")

            if '库存数据' in merged_file.sheet_names:
                inventory_df = pd.read_excel(output_path, sheet_name='库存数据', engine='openpyxl')
                non_numeric = check_numeric_columns(inventory_df, "库存数据", 6, 25)
                all_non_numeric_cells.extend(non_numeric)
                if non_numeric:
                    logger.warning(f"库存数据中发现 {len(non_numeric)} 个非数字单元格")

        except Exception as e:
            logger.error(f"检查数字格式时出错: {e}")

        logger.info(f"操作完成！新文件已保存为: {output_path}")
        return True, output_path, duplicate_info, all_non_numeric_cells, file1_time_str, file2_time_str

    except Exception as e:
        error_msg = f"程序执行出错: {e}"
        logger.error(error_msg)
        return False, error_msg, None, None, "未知", "未知"


def process_with_duplicates(output_path, duplicate_info, connection_string, container_name, output_filename):
    """
    处理包含重复物料的情况
    """
    logger.info(f"发现 {len(duplicate_info)} 个重复物料，自动删除重复项...")

    for item in duplicate_info:
        logger.info(f"  - {item['物料名称']} (出现{item['次数']}次) - 采购负责人: {item['采购负责人']}")

    # 自动删除重复项
    logger.info("自动删除重复项并继续上传...")

    try:
        # 读取文件，删除重复项
        inventory_df = pd.read_excel(output_path, sheet_name='库存数据', engine='openpyxl')
        purchase_df = pd.read_excel(output_path, sheet_name='采购数据', engine='openpyxl')

        # 删除重复项，保留第一次出现
        cleaned_inventory, deleted_count = remove_duplicate_material_rows(inventory_df, duplicate_info)

        # 重新保存文件
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            cleaned_inventory.to_excel(writer, sheet_name='库存数据', index=False)
            purchase_df.to_excel(writer, sheet_name='采购数据', index=False)

        logger.info(f"已删除 {deleted_count} 行重复数据")

        # 上传到Azure
        upload_success, result = upload_to_azure_blob(
            local_file_path=output_path,
            connection_string=connection_string,
            container_name=container_name,
            blob_name=output_filename
        )

        return upload_success, result, deleted_count

    except Exception as e:
        error_msg = f"处理重复数据时出错: {str(e)}"
        logger.error(f"{error_msg}")
        return False, error_msg, 0


def main():
    """主函数 - 固定处理两个特定文件并自动上传到Azure"""

    # 设置文件夹路径
    folder_path = r"C:\Users\Administrator\Downloads"  # 修改为你的文件夹路径

    # 固定文件名
    file1_name = "原料一科采购日报表.xlsx"  # 只使用文件名，不包含路径
    file2_name = "原料二科采购日报表.xlsx"  # 只使用文件名，不包含路径
    output_filename = "mergedexcel.xlsx"

    # 获取钉钉配置
    webhook, secret = get_dingtalk_config()

    # 发送程序启动消息
    start_msg = f"开始处理文件: {file1_name}, {file2_name}"
    logger.info("=" * 50)
    logger.info("Excel合并程序开始运行")
    logger.info("=" * 50)

    # 发送钉钉消息
    send_status_update("Excel文件合并", "start", start_msg, webhook=webhook, secret=secret)

    start_time = time.time()

    try:
        # 运行合并程序
        logger.info("开始合并Excel文件...")
        success, output_path, duplicate_info, non_numeric_cells, file1_time_str, file2_time_str = merge_excel_sheets(
            folder_path=folder_path,
            file1_name=file1_name,
            file2_name=file2_name,
            output_filename=output_filename
        )

        if not success:
            # 合并失败，发送错误消息
            error_details = output_path  # 这里output_path实际上是错误信息
            error_msg = f"合并过程中发生错误: {error_details}"
            send_status_update("Excel文件合并", "error", error_msg, webhook=webhook, secret=secret)
            logger.error(error_msg)
            return

        # 合并成功，准备消息
        end_time = time.time()
        duration = round(end_time - start_time, 2)

        # 检查文件大小
        file_size = os.path.getsize(output_path) if os.path.exists(output_path) else 0
        file_size_mb = round(file_size / (1024 * 1024), 2) if file_size > 0 else 0

        # 读取合并后的文件统计信息
        try:
            merged_file = pd.ExcelFile(output_path, engine='openpyxl')
            inventory_count = 0
            purchase_count = 0

            if '库存数据' in merged_file.sheet_names:
                inventory_df = pd.read_excel(output_path, sheet_name='库存数据', engine='openpyxl')
                inventory_count = len(inventory_df)

            if '采购数据' in merged_file.sheet_names:
                purchase_df = pd.read_excel(output_path, sheet_name='采购数据', engine='openpyxl')
                purchase_count = len(purchase_df)

        except Exception as e:
            logger.error(f"读取合并文件统计信息时出错: {e}")
            inventory_count = 0
            purchase_count = 0

        # 构建基础消息
        message = f"**处理时间**: {duration}秒\n\n"
        message += f"**输出文件**: {output_filename} ({file_size_mb}MB)\n\n"
        message += f"**库存数据**: {inventory_count} 行\n\n"
        message += f"**采购数据**: {purchase_count} 行\n\n"
        message += f"**文件刷新时间**:\n- 原料一科: {file1_time_str}\n- 原料二科: {file2_time_str}\n\n"

        if non_numeric_cells:
            message += f"**发现 {len(non_numeric_cells)} 个非数字单元格**\n"
            for i, cell in enumerate(non_numeric_cells[:3]):  # 只显示前3个
                message += f"- 行 {cell['行号']}, 列 '{cell['列名']}': '{cell['值']}'\n"
            if len(non_numeric_cells) > 3:
                message += f"... 等 {len(non_numeric_cells)} 个非数字单元格\n"

        # 检查是否有重复物料
        if duplicate_info:
            logger.info(f"发现 {len(duplicate_info)} 个重复物料，发送通知并处理...")

            # 发送重复物料通知
            duplicate_message = f"发现 {len(duplicate_info)} 个重复物料名称"
            send_status_update("Excel文件合并", "duplicates", duplicate_message, duplicate_info, webhook=webhook,
                               secret=secret)

            # 获取Azure配置
            connection_string, container_name = get_azure_config()

            if not connection_string:
                error_msg = "未配置Azure连接字符串，无法上传"
                send_status_update("Excel文件合并", "warning", f"{message}\n\n{error_msg}", webhook=webhook,
                                   secret=secret)
                logger.error(error_msg)
                return

            # 处理重复物料
            logger.info("处理重复物料...")
            upload_success, upload_result, deleted_count = process_with_duplicates(
                output_path, duplicate_info, connection_string, container_name, output_filename
            )

            if upload_success:
                success_message = f"{message}\n**重复物料处理**: 已删除 {deleted_count} 行重复数据\n\n"
                success_message += f"**上传状态**: 成功\n\n"
                success_message += f"**文件位置**: {upload_result}"

                send_status_update("Excel合并与上传完成", "success", success_message, webhook=webhook, secret=secret)
                logger.info("文件处理并上传成功")
            else:
                error_message = f"{message}\n**重复物料处理**: 处理失败\n**错误信息**: {upload_result}"
                send_status_update("Excel合并完成但上传失败", "error", error_message, webhook=webhook, secret=secret)
                logger.error(f"文件处理或上传失败: {upload_result}")

        else:
            # 没有重复物料，直接上传
            logger.info("没有发现重复物料，直接上传到Azure...")

            # 获取Azure配置
            connection_string, container_name = get_azure_config()

            if not connection_string:
                error_msg = "未配置Azure连接字符串，无法上传"
                send_status_update("Excel文件合并", "warning", f"{message}\n\n{error_msg}", webhook=webhook,
                                   secret=secret)
                logger.error(error_msg)
                return

            # 上传到Azure
            logger.info("开始上传到Azure Blob Storage...")
            upload_success, upload_result = upload_to_azure_blob(
                local_file_path=output_path,
                connection_string=connection_string,
                container_name=container_name,
                blob_name=output_filename
            )

            if upload_success:
                success_message = f"{message}\n**上传状态**: 成功\n\n**文件位置**: {upload_result}"
                send_status_update("Excel合并与上传完成", "success", success_message, webhook=webhook, secret=secret)
                logger.info("文件上传成功")
            else:
                error_message = f"{message}\n**上传状态**: 失败\n**错误信息**: {upload_result}"
                send_status_update("Excel合并完成但上传失败", "error", error_message, webhook=webhook, secret=secret)
                logger.error(f"文件上传失败: {upload_result}")

    except Exception as e:
        error_msg = f"程序执行过程中发生错误: {str(e)}"
        logger.error(f"{error_msg}")
        send_status_update("Excel合并程序异常", "error", error_msg, webhook=webhook, secret=secret)

    logger.info("=" * 50)
    logger.info("Excel合并程序运行结束")
    logger.info("=" * 50)


if __name__ == "__main__":
    # 安装所需库的提示
    try:
        import pandas as pd
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill
        from azure.storage.blob import BlobServiceClient
        import requests
    except ImportError as e:
        print("请先安装所需库:")
        print("pip install pandas openpyxl azure-storage-blob requests")
        exit(1)

    # 运行主程序
    main()